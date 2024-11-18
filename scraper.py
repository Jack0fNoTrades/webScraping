from selenium import webdriver
import openpyxl
from openpyxl.workbook import Workbook
import json
import pandas as pd


def bizbuysellExcel():
    try:
        wb = openpyxl.load_workbook("Businesses For Sale - BizBuySell.xlsx")
    except FileNotFoundError:
        wb = Workbook()
    finally:
        ws = wb.active
        ws['A1'] = "type"
        ws['B1'] = "name"
        ws['C1'] = "logo"
        ws['D1'] = "image"
        ws['E1'] = "description"
        ws['F1'] = "url"
        ws['G1'] = "price"
        ws['H1'] = "currency"
        ws['I1'] = "city"
        ws['J1'] = "state/country"

        row = 1
        startFound = False

        while ws['A' + str(row)].value is not None:
            row += 1

        row -= 1

        with open("Businesses For Sale - BizBuySell.txt", "r") as datafile:
            contents = datafile.readlines()

        for line in contents:
            if "about" in line and line.endswith("[\n"):
                startFound = True
            elif len(line) == 3 and line[1] == ']' and startFound:
                break
            elif startFound:
                if "@type" in line and ("Product" in line or "Person" in line):
                    row += 1
                    line = line.strip()
                    ws['A' + str(row)] = line[10:len(line) - 2]
                elif "name" in line:
                    line = line.strip()
                    ws['B' + str(row)] = line[9:len(line) - 2]
                elif "logo" in line:
                    line = line.strip()
                    ws['C' + str(row)] = line[9:len(line) - 2]
                elif "image" in line:
                    line = line.strip()
                    ws['D' + str(row)] = line[10:len(line) - 2]
                elif "description" in line:
                    line = line.strip()
                    ws['E' + str(row)] = line[16:len(line) - 2]
                elif "url" in line:
                    line = line.strip()
                    listing_url = "https://www.bizbuysell.com" + line[8:len(line) - 2]
                    ws['F' + str(row)] = listing_url
                elif "price" in line and "priceSpecification" not in line and "Currency" not in line:
                    line = line.strip()
                    ws['G' + str(row)] = line[9:len(line) - 1]
                elif "priceCurrency" in line:
                    line = line.strip()
                    ws['H' + str(row)] = line[18:len(line) - 2]
                elif "addressLocality" in line:
                    line = line.strip()
                    if line[19:len(line) - 1] != "null":
                        ws['I' + str(row)] = line[20:len(line) - 2]
                elif "addressRegion" in line:
                    line = line.strip()
                    ws['J' + str(row)] = line[18:len(line) - 1]

        wb.save("Businesses For Sale - BizBuySell.xlsx")


def flippaExcel():
    outfile = "Online Business For Sale | 5,000 Listings | Flippa.xlsx"
    df_list = []

    try:
        existingData = pd.read_excel(outfile)
        df_list.append(existingData)
    except FileNotFoundError:
        pass
    finally:

        data = dict()

        with open("Online Business For Sale | 5,000 Listings | Flippa.txt", "r") as file:
            lines = file.readlines()

        for line in lines:
            if "const STATE = {" in line:
                datastr = line.strip()[14:len(line) - 4]
                data = json.loads(datastr)
                break

        newData = pd.DataFrame(data=data["results"], index=[i for i in range(25)])
        df_list.append(newData)
        df = pd.concat(df_list)

        df.to_excel(outfile)


with open("urls.txt", "r") as websites:
    urls = websites.readlines()

options = webdriver.ChromeOptions()
options.add_argument("--headless=new")

for url in urls:
    driver = webdriver.Chrome(options=options)
    driver.get(url)
    filename = driver.title + ".txt"
    with open(filename, "a") as file:
        file.write(driver.page_source)
        file.close()
    driver.quit()

bizbuysellExcel()
flippaExcel()
